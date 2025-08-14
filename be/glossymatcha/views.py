from .models import Staff, Suppliers, DailySales, Sales, Inquiries, Products, WorkRecord, YearlySales, DailyPassword
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import date, datetime, timedelta
from django.db.models import Sum, Q
from django.db import models
from django.contrib import messages
from django.urls import reverse_lazy
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_protect
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView
from .serializers import InquirySerializer, ProductListSerializer, ProductDetailSerializer
from .forms import StaffForm, WorkRecordForm, SuppliersForm, DailySalesForm, SalesForm, YearlySalesForm, DailyPasswordForm, DailyPasswordManagementForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils.dataframe import dataframe_to_rows
import json, calendar

class DashboardView(LoginRequiredMixin, TemplateView):
    """
    대시보드 메인 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 대시보드 페이지를 렌더링
    - 통계 데이터: 직원 수, 공급업체 수, 오늘 매출, 이번 달 매출, 최근 매출 현황, 직원 목록, 최근 문의 현황
    """
    template_name = 'glossymatcha/dashboard.html'

    def dispatch(self, request, *args, **kwargs):
        """
        요청을 처리하기 전에 인증 및 세션 검증
        - 사용자가 인증되지 않은 경우 로그인 페이지로 리다이렉트
        - 일일 비밀번호 검증이 완료되지 않은 경우 일일 비밀번호 확인 페이지로 리다이렉트
        """
        if not request.user.is_authenticated:
            return redirect('login')
        
        if not request.session.get('daily_password_verified'):
            return redirect('daily_password_check')
        
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()

        # 통계 데이터 수집
        active_staff_count = Staff.objects.filter(
            Q(resignation_date__isnull=True) | Q(resignation_date__gt=today)
        ).count()
        active_suppliers_count = Suppliers.objects.filter(is_active=True).count()

        # 오늘 매출 집계
        today_daily_sales = DailySales.objects.filter(date=today).first()
        today_sales = today_daily_sales.total_sales if today_daily_sales else 0

        # 이번 달 매출 집계 (년 | 월 형식)
        current_month_sales = Sales.objects.filter(
            year=today.year,
            month=today.month
        ).first()
        monthly_sales = current_month_sales.total_sales if current_month_sales else 0
        
        # 이번 연도 매출 집계
        current_year_sales = YearlySales.objects.filter(year=today.year).first()
        yearly_sales = current_year_sales.total_sales if current_year_sales else 0

        # 최근 매출 현황 (최근 5건 - 일별 매출)
        recent_sales = DailySales.objects.all()[:5]

        # 직원 목록 (재직중인 최근 5명)
        staff_list = Staff.objects.filter(
            Q(resignation_date__isnull=True) | Q(resignation_date__gt=today)
        )[:5]

        # 최근 문의 현황 (최근 5건)
        recent_inquiries = Inquiries.objects.all()[:5]
        total_inquiries = Inquiries.objects.count()

        # chart data creation
        chart_data = self.get_chart_data()
        sales_analysis = self.get_sales_analysis()
        
        context.update({
            'today': today,
            'active_staff_count': active_staff_count,
            'active_suppliers_count': active_suppliers_count,
            'today_sales': today_sales,
            'monthly_sales': monthly_sales,
            'yearly_sales': yearly_sales,
            'recent_sales': recent_sales,
            'staff_list': staff_list,
            'recent_inquiries': recent_inquiries,
            'total_inquiries': total_inquiries,
            'chart_data': json.dumps(chart_data),
            'sales_analysis': sales_analysis,
        })

        return context
    
    def get_chart_data(self):
        """차트 데이터 생성 - 최근 12개월간의 월별 매출 데이터"""
        today = date.today()
        # 최근 12개월 데이터
        months_data = []
        labels = []

        for i in range(11, -1, -1):
            """12개월 전부터 현재까지의 월별 매출 데이터 생성"""
            # 월을 직접 계산
            year = today.year
            month = today.month - i

            # 월이 0 이하가 되면 이전 년도로
            while month <= 0:
                month += 12
                year -= 1

            # 해당 월의 매출 데이터 조회
            monthly_sales = Sales.objects.filter(year=year, month=month).first()
            sales_amount = monthly_sales.total_sales if monthly_sales else 0
            cost_amount = monthly_sales.total_cost if monthly_sales else 0
            profit_amount = monthly_sales.gross_profit if monthly_sales else 0

            labels.append(f"{year}년 {month}월")
            months_data.append({
                'sales': float(sales_amount),
                'cost': float(cost_amount),
                'profit': float(profit_amount)
            })
    
        # 최근 30일 일별 매출 데이터
        daily_data = []
        daily_labels = []
        for i in range(29, -1, -1):
            """30일 전부터 현재까지의 일별 매출 데이터 생성"""
            target_date = today - timedelta(days=i)
            daily_sales = DailySales.objects.filter(date=target_date).first()
            sales_amount = daily_sales.total_sales if daily_sales else 0

            daily_labels.append(target_date.strftime('%m/%d'))
            daily_data.append(float(sales_amount))

        return {
            'monthly': {
                'labels': labels,
                'sales': [item['sales'] for item in months_data],
                'cost': [item['cost'] for item in months_data],
                'profits': [item['profit'] for item in months_data]
            },
            'daily': {
                'labels': daily_labels,
                'sales': daily_data
            }
        }
    
    def get_sales_analysis(self):
        """매출 분석 데이터 생성"""
        today = date.today()
        
        # 이번 달
        current_year = today.year
        current_month_num = today.month

        # 저번 달 매출 계산
        previous_month_num = current_month_num - 1
        previous_year = current_year
        if previous_month_num <= 0:
            previous_month_num = 12
            previous_year -= 1
        
        # 작년 동월
        last_year = current_year - 1
        same_month_last_year = current_month_num

        # 이번 달 매출
        current_sales = Sales.objects.filter(year=current_year, month=current_month_num).first()
        current_amount = current_sales.total_sales if current_sales else 0
        current_cost = current_sales.total_cost if current_sales else 0
        current_profit = current_sales.gross_profit if current_sales else 0

        # 저번 달 매출
        previous_sales = Sales.objects.filter(year=previous_year, month=previous_month_num).first()
        previous_amount = previous_sales.total_sales if previous_sales else 0

        # 작년 동월 매출
        last_year_sales = Sales.objects.filter(year=last_year, month=same_month_last_year).first()
        last_year_amount = last_year_sales.total_sales if last_year_sales else 0

        # 매출 증감 계산
        mom_change = ((current_amount - previous_amount) / previous_amount * 100) if previous_amount > 0 else 0
        yoy_change = ((current_amount - last_year_amount) / last_year_amount * 100) if last_year_amount > 0 else 0

        # 수익률 계산
        profit_margin = (current_profit / current_amount * 100) if current_amount > 0 else 0
        cost_ratio = (current_cost / current_amount * 100) if current_amount > 0 else 0

        return {
            'current_month': current_amount,
            'current_profit': current_profit,
            'current_cost': current_cost,
            'mom_change': round(mom_change, 1),
            'yoy_change': round(yoy_change, 1),
            'profit_margin': round(profit_margin, 1),
            'cost_ratio': round(cost_ratio, 1),
        }

# API Views for Inquiries
class CreateInquiryView(generics.CreateAPIView):
    """
    문의 생성 API
    - 모든 사용자 접근 가능
    - 문의 내용을 저장하고 응답
    Request body 예시:
    {
        "name": "문의자 이름 (필수)",
        "email": "문의자 이메일 (필수)",
        "inquiry_type": "문의 유형 (general/product/other, 기본값: general)",
        "message": "문의 내용 (필수)"
    }
    Response:
    - 201: 문의 생성 성공
    - 400: 입력 데이터 검증 실패
    """
    queryset = Inquiries.objects.all()
    serializer_class = InquirySerializer
    permission_classes = [AllowAny]

class InquiryListView(generics.ListAPIView):
    """
    문의 목록 API
    - 관리자만 접근 가능
    - 모든 문의를 조회
    
    Response body 예시:
    [
        {
            "id": 1,
            "name": "문의자 이름",
            "email": "문의자 이메일",
            "inquiry_type": "general",
            "message": "문의 내용",
            "created_at": "2023-10-01T12:00:00Z"
        },
        ...
    ]
    Response:
    - 200: 문의 목록 조회 성공
    - 403: 권한 없음 (관리자만 접근 가능)
    """
    queryset = Inquiries.objects.all()
    serializer_class = InquirySerializer
    permission_classes = [IsAdminUser]

class InquiryDetailView(generics.RetrieveAPIView):
    """
    문의 상세 조회 API
    - 관리자만 접근 가능
    - 특정 문의의 상세 정보를 조회
    Response body 예시:
    {
        "id": 1,
        "name": "문의자 이름",
        "email": "문의자 이메일",
        "inquiry_type": "general",
        "message": "문의 내용",
        "created_at": "2023-10-01T12:00:00Z"
    }
    Response:
    - 200: 문의 상세 조회 성공
    - 404: 해당 문의가 존재하지 않음
    - 403: 권한 없음 (관리자만 접근 가능)
    """
    queryset = Inquiries.objects.all()
    serializer_class = InquirySerializer
    permission_classes = [IsAdminUser]

# TemplateView for Inquiry Form
class InquiriesListView(LoginRequiredMixin, ListView):
    """
    문의 목록 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 문의 목록 페이지를 렌더링
    """
    model = Inquiries
    template_name = 'glossymatcha/inquiries/list.html'
    context_object_name = 'inquiries_list'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 유형별 문의 수 집계
        queryset = self.get_queryset()
        context['general_count'] = queryset.filter(inquiry_type='general').count()
        context['product_count'] = queryset.filter(inquiry_type='product').count()
        context['other_count'] = queryset.filter(inquiry_type='other').count()

        return context

class InquiriesDetailView(LoginRequiredMixin, DetailView):
    """
    문의 상세 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 문의 상세 페이지를 렌더링
    """
    model = Inquiries
    template_name = 'glossymatcha/inquiries/detail.html'
    context_object_name = 'inquiry'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # 같은 이메일의 다른 문의들 조회 (최근 5건, 본인 문의 포함)
        related_inquiries = Inquiries.objects.filter(
            email=self.object.email
        ).order_by('-created_at')[:5]
        context['related_inquiries'] = related_inquiries

        return context

class InquiriesDeleteView(LoginRequiredMixin, DeleteView):
    """
    문의 삭제 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 문의 삭제 페이지를 렌더링
    """
    model = Inquiries
    template_name = 'glossymatcha/inquiries/delete.html'
    success_url = reverse_lazy('inquiries_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(request, f'{self.object.name}님의 문의가 삭제되었습니다.')
        return super().delete(request, *args, **kwargs)
    
# API Views for Products
class ProductListView(generics.ListAPIView):
    """
    제품 목록 API
    - 모든 사용자 접근 가능
    - 제품 목록을 조회하고 응답
    Request query parameters:
    - lang: 언어 코드 (ko/en, 기본값: ko)
    Response body 예시:
    {
        "success": true,
        "language": "ko",
        "count": 10,
        "results": [
            {
                "id": 1,
                "name": "제품 이름",
                "description": "제품 설명",
                ...
            },
            ...
        ]
    }
    Response:
    - 200: 제품 목록 조회 성공
    - 400: 잘못된 언어 코드
    """
    queryset = Products.objects.all()
    serializer_class = ProductListSerializer
    permission_classes = [AllowAny]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        # url 파라미터에서 언어 정보 추출(기본값 : 'ko')
        language = self.request.query_params.get('lang', 'ko')
        if language not in ['ko', 'en']:
            language = 'ko'
        context['language'] = language
        return context
    
    def list(self, request, *args, **kwargs):
        """
        GET 요청에 대한 처리
        - 제품 목록을 조회하고 응답
        - 언어 파라미터에 따라 제품 이름과 설명을 번역
        """
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'success': True,
            'language': self.get_serializer_context()['language'],
            'count': len(serializer.data),
            'results': serializer.data
        })

class ProductDetailView(generics.RetrieveAPIView):
    """
    제품 상세 조회 API
    - 모든 사용자 접근 가능
    - 특정 제품의 상세 정보를 조회
    Request path parameters:
    - pk: 제품 ID
    Response body 예시:
    {
        "success": true,
        "language": "ko",
        "result": {
            "id": 1,
            "name": "제품 이름",
            "description": "제품 설명",
            "images": [
                {
                    "id": 1,
                    "url": "https://example.com/image1.jpg"
                },
                ...
            ],
            "specifications": [
                {
                    "id": 1,
                    "key": "사양 키",
                    "value": "사양 값"
                },
                ...
            ]
        }
    }
    Response:
    - 200: 제품 상세 조회 성공
    - 404: 해당 제품이 존재하지 않음
    """
    queryset = Products.objects.all()
    serializer_class = ProductDetailSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return super().get_queryset().prefetch_related('images', 'specifications')
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        # url 파라미터에서 언어 정보 추출(기본값 : 'ko')
        language = self.request.query_params.get('lang', 'ko')
        if language not in ['ko', 'en']:
            language = 'ko'
        context['language'] = language
        return context
    
    def retrieve(self, request, *args, **kwargs):
        """
        GET 요청에 대한 처리
        - 특정 제품의 상세 정보를 조회하고 응답
        - 언어 파라미터에 따라 제품 이름과 설명을 번역
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({
            'success': True,
            'language': self.get_serializer_context()['language'],
            'result': serializer.data
        })
    
class StaffListView(LoginRequiredMixin, ListView):
    """
    직원 목록 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 직원 목록 페이지를 렌더링
    """
    model = Staff
    template_name = 'glossymatcha/staff/list.html'
    context_object_name = 'staff_list'
    
    def get_queryset(self):
        """
        직원 목록을 정렬하여 반환
        - 재직 중인 직원은 먼저, 퇴사한 직원은 나중에
        - 정직원은 먼저, 파트직원은 나중에
        - hire_date 기준으로 정렬
        """
        return Staff.objects.annotate(
            # 정렬 우선순위: 재직 > 퇴사
            status_order=models.Case(
                models.When(resignation_date__isnull=True, then=models.Value(0)),  # 재직중
                default=models.Value(1),  # 퇴사
                output_field=models.IntegerField(),
            ),
            # 직종 정렬 우선순위: 정직원 > 파트직원 > 퇴사
            employee_type_order=models.Case(
                models.When(employee_type='full_time', then=models.Value(0)),  # 정직원
                models.When(employee_type='part_time', then=models.Value(1)),   # 파트직원
                models.When(employee_type='resigned', then=models.Value(2)),    # 퇴사
                default=models.Value(3),
                output_field=models.IntegerField(),
            )
        ).order_by('status_order', 'employee_type_order', 'hire_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        staff_list = context['staff_list']
        
        # 재직 중인 직원만 필터링
        active_staff = [staff for staff in staff_list if staff.is_active]
        
        # 재직 중인 직원 수 계산
        active_staff_count = len(active_staff)
        active_full_time_count = len([staff for staff in active_staff if staff.employee_type == 'full_time'])
        active_part_time_count = len([staff for staff in active_staff if staff.employee_type == 'part_time'])
        
        context.update({
            'active_staff_count': active_staff_count,
            'active_full_time_count': active_full_time_count,
            'active_part_time_count': active_part_time_count,
        })
        
        return context

class StaffCreateView(LoginRequiredMixin, CreateView):
    """
    직원 생성 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 직원 생성 페이지를 렌더링
    """
    model = Staff
    form_class = StaffForm
    template_name = 'glossymatcha/staff/create.html'
    success_url = reverse_lazy('staff_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.name} 직원이 성공적으로 등록되었습니다.')
        return response
    
class StaffDetailView(LoginRequiredMixin, DetailView):
    """
    직원 상세 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 직원 상세 페이지를 렌더링
    """
    model = Staff
    template_name = 'glossymatcha/staff/detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 최근 급여 기록
        recent_work_records = WorkRecord.objects.filter(staff=self.object).order_by('-created_at')[:10]
        context['recent_work_records'] = recent_work_records            
        
        return context
    
class StaffUpdateView(LoginRequiredMixin, UpdateView):
    """
    직원 수정 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 직원 수정 페이지를 렌더링
    """
    model = Staff
    form_class = StaffForm
    template_name = 'glossymatcha/staff/create.html'

    def get_success_url(self):
        return reverse_lazy('staff_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.name} 직원 정보가 수정되었습니다.')
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['staff'] = self.object
        return context
    
class StaffDeleteView(LoginRequiredMixin, DeleteView):
    """
    직원 삭제 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 직원 삭제 페이지를 렌더링
    """
    model = Staff
    template_name = 'glossymatcha/staff/delete.html'
    success_url = reverse_lazy('staff_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(request, f'{self.object.name} 직원이 삭제되었습니다.')
        return super().delete(request, *args, **kwargs)
    
class WorkRecordCreateView(LoginRequiredMixin, CreateView):
    """
    직원 근무 기록 생성 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 근무 기록 생성 페이지를 렌더링
    """
    model = WorkRecord
    form_class = WorkRecordForm
    template_name = 'glossymatcha/staff/work_record.html'

    def get_success_url(self):
        return reverse_lazy('staff_detail', kwargs={'pk': self.object.staff.pk})

    def get_initial(self):
        initial = super().get_initial()
        # URL 파라미터에서 staff_id를 가져와서 초기값으로 설정
        staff_id = self.request.GET.get('staff_id')
        if staff_id:
            try:
                staff = Staff.objects.get(id=staff_id)
                initial['staff'] = staff
            except Staff.DoesNotExist:
                pass
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # staff_id를 템플릿에 전달
        staff_id = self.request.GET.get('staff_id')
        if staff_id:
            try:
                context['selected_staff'] = Staff.objects.get(id=staff_id)
            except Staff.DoesNotExist:
                pass
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.staff.name}의 근무 기록이 등록되었습니다.')
        return response

class WorkRecordUpdateView(LoginRequiredMixin, UpdateView):
    """
    직원 근무 기록 수정 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 근무 기록 수정 페이지를 렌더링
    """
    model = WorkRecord
    form_class = WorkRecordForm
    template_name = 'glossymatcha/staff/work_record.html'

    def get_success_url(self):
        return reverse_lazy('staff_detail', kwargs={'pk': self.object.staff.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_update'] = True
        context['selected_staff'] = self.object.staff
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.staff.name}의 근무 기록이 수정되었습니다.')
        return response

class WorkRecordDeleteView(LoginRequiredMixin, DeleteView):
    """
    직원 근무 기록 삭제 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 근무 기록 삭제 페이지를 렌더링
    """
    model = WorkRecord
    template_name = 'glossymatcha/staff/work_record_delete.html'

    def get_success_url(self):
        return reverse_lazy('staff_detail', kwargs={'pk': self.object.staff.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['staff'] = self.object.staff
        return context

    def delete(self, request, *args, **kwargs):
        staff_name = self.get_object().staff.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'{staff_name}의 근무 기록이 삭제되었습니다.')
        return response
    
# Django Template Views for suppliers management
class SuppliersListView(LoginRequiredMixin, ListView):
    """
    공급업체 목록 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 공급업체 목록 페이지를 렌더링
    """
    model = Suppliers
    template_name = 'glossymatcha/suppliers/list.html'
    context_object_name = 'suppliers_list'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        suppliers_list = context['suppliers_list']
        
        # 거래처 통계 계산
        active_count = suppliers_list.filter(is_active=True).count()
        cash_count = suppliers_list.filter(payment_method='cash', is_active=True).count()
        card_count = suppliers_list.filter(payment_method='card', is_active=True).count()
        transfer_count = suppliers_list.filter(payment_method='transfer', is_active=True).count()
        credit_count = suppliers_list.filter(payment_method='credit', is_active=True).count()
        
        context.update({
            'active_count': active_count,
            'cash_count': cash_count,
            'card_count': card_count,
            'transfer_count': transfer_count,
            'credit_count': credit_count,
        })
        
        return context

class SuppliersCreateView(LoginRequiredMixin, CreateView):
    """
    공급업체 생성 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 공급업체 생성 페이지를 렌더링
    """
    model = Suppliers
    form_class = SuppliersForm
    template_name = 'glossymatcha/suppliers/create.html'
    success_url = reverse_lazy('suppliers_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.name} 거래처가 등록되었습니다.')
        return response
    
class SuppliersDetailView(LoginRequiredMixin, DetailView):
    """
    공급업체 상세 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 공급업체 상세 페이지를 렌더링
    """
    model = Suppliers
    template_name = 'glossymatcha/suppliers/detail.html'
    context_object_name = 'supplier'

class SuppliersUpdateView(LoginRequiredMixin, UpdateView):
    """
    공급업체 수정 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 공급업체 수정 페이지를 렌더링
    """
    model = Suppliers
    form_class = SuppliersForm
    template_name = 'glossymatcha/suppliers/create.html'
    
    def get_success_url(self):
        return reverse_lazy('suppliers_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.name} 거래처 정보가 수정되었습니다.')
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['supplier'] = self.object
        return context
    
class SuppliersDeleteView(LoginRequiredMixin, DeleteView):
    """
    공급업체 삭제 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 공급업체 삭제 페이지를 렌더링
    """
    model = Suppliers
    template_name = 'glossymatcha/suppliers/delete.html'
    success_url = reverse_lazy('suppliers_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(request, f'{self.object.name} 거래처가 삭제되었습니다.')
        return super().delete(request, *args, **kwargs)

# Django Template Views for daily sales management
class DailySalesListView(LoginRequiredMixin, ListView):
    """
    일별 매출 목록 조회
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 일별 매출 목록 페이지를 렌더링
    """
    model = DailySales
    template_name = 'glossymatcha/daily_sales/list.html'
    context_object_name = 'daily_sales_list'


class DailySalesCreateView(LoginRequiredMixin, CreateView):
    """
    일별 매출 등록
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 일별 매출 등록 페이지를 렌더링"""
    model = DailySales
    form_class = DailySalesForm
    template_name = 'glossymatcha/daily_sales/create.html'
    success_url = reverse_lazy('daily_sales_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.date} 일별 매출이 등록되었습니다.')
        return response


class DailySalesDetailView(LoginRequiredMixin, DetailView):
    """
    일별 매출 상세 정보
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 일별 매출 상세 페이지를 렌더링
    """
    model = DailySales
    template_name = 'glossymatcha/daily_sales/detail.html'
    context_object_name = 'daily_sale'


class DailySalesUpdateView(LoginRequiredMixin, UpdateView):
    """
    일별 매출 정보 수정
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 일별 매출 수정 페이지를 렌더링
    """
    model = DailySales
    form_class = DailySalesForm
    template_name = 'glossymatcha/daily_sales/create.html'
    
    def get_success_url(self):
        return reverse_lazy('daily_sales_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.date} 일별 매출 정보가 수정되었습니다.')
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['daily_sale'] = self.object
        return context


class DailySalesDeleteView(LoginRequiredMixin, DeleteView):
    """
    일별 매출 삭제
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 일별 매출 삭제 페이지를 렌더링
    """
    model = DailySales
    template_name = 'glossymatcha/daily_sales/delete.html'
    success_url = reverse_lazy('daily_sales_list')
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(request, f'{self.object.date} 일별 매출이 삭제되었습니다.')
        return super().delete(request, *args, **kwargs)

# Django Template Views for monthly sales management
class SalesListView(LoginRequiredMixin, ListView):
    """
    월별 매출 목록 조회
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 월별 매출 목록 페이지를 렌더링
    """
    model = Sales
    template_name = 'glossymatcha/sales/list.html'
    context_object_name = 'sales_list'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 총합 계산
        queryset = self.get_queryset()
        total_sales_sum = sum(sale.total_sales or 0 for sale in queryset)
        total_cost_sum = sum(sale.total_cost or 0 for sale in queryset)
        total_profit_sum = sum(sale.gross_profit or 0 for sale in queryset)
        
        context.update({
            'total_sales_sum': total_sales_sum,
            'total_cost_sum': total_cost_sum,
            'total_profit_sum': total_profit_sum,
        })
        
        return context

class SalesCreateView(LoginRequiredMixin, CreateView):
    """
    월별 매출 등록
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 월별 매출 등록 페이지를 렌더링
    """
    model = Sales
    form_class = SalesForm
    template_name = 'glossymatcha/sales/create.html'
    success_url = reverse_lazy('sales_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.year}년 {self.object.month}월 매출이 등록되었습니다.')
        return response
    
class SalesDetailView(LoginRequiredMixin, DetailView):
    """
    월별 매출 상세 정보
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 월별 매출 상세 페이지를 렌더링
    """
    model = Sales
    template_name = 'glossymatcha/sales/detail.html'
    context_object_name = 'sale'

class SalesUpdateView(LoginRequiredMixin, UpdateView):
    """
    월별 매출 정보 수정
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 월별 매출 수정 페이지를 렌더링
    """
    model = Sales
    form_class = SalesForm
    template_name = 'glossymatcha/sales/create.html'

    def get_success_url(self):
        return reverse_lazy('sales_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.year}년 {self.object.month}월 매출 정보가 수정되었습니다.')
        return response
    
    def get_context_data(self, **kwargs):
        """
        월별 매출 수정 페이지의 컨텍스트 데이터 추가
        - 현재 매출 객체를 컨텍스트에 추가
        """
        context = super().get_context_data(**kwargs)
        context['sale'] = self.object
        return context
    
class SalesDeleteView(LoginRequiredMixin, DeleteView):
    """
    월별 매출 삭제
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 월별 매출 삭제 페이지를 렌더링
    """
    model = Sales
    template_name = 'glossymatcha/sales/delete.html'
    success_url = reverse_lazy('sales_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(request, f'{self.object.year}년 {self.object.month}월 매출이 삭제되었습니다.')
        return super().delete(request, *args, **kwargs)
    
class MonthlySalesExcelExportView(LoginRequiredMixin, TemplateView):
    """
    월별 매출 현황 엑셀 파일 다운로드
    - 로그인한 사용자만 접근 가능
    - 월별 매출 데이터를 엑셀 파일로 생성하여 다운로드"""
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = '월별 매출 현황'

        # 헤더 스타일 설정
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # 콤마 포맷팅 설정
        number_style = NamedStyle(name='number_comma')
        number_style.number_format = '#,##0'

        # 퍼센트 포맷 설정
        percent_style = NamedStyle(name='percent')
        percent_style.number_format = '0.0%'

        # 헤더 작성
        headers = [
            '연도', '월', '오프라인 매출', '기타 매출', '총 매출',
            '재료비', '노무비', '경비-소모품', '경비-기타', '총 원가', '매출 총이익', '수익률(%)'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 데이터 작성
        sales_data = Sales.objects.all().order_by('-year', '-month')

        for row, sale in enumerate(sales_data, 2):
            profit_margin = (sale.gross_profit / sale.total_sales) if sale.total_sales > 0 else 0

            # 기본 데이터 작성
            ws.cell(row=row, column=1, value=sale.year)
            ws.cell(row=row, column=2, value=sale.month)

            # 금액 데이터 작성
            money_cells = [
                (3, int(sale.offline_sales)),
                (4, int(sale.other_sales)),
                (5, int(sale.total_sales)),
                (6, int(sale.material_cost)),
                (7, int(sale.labor_cost)),
                (8, int(sale.supplies_expense)),
                (9, int(sale.other_expense)),
                (10, int(sale.total_cost)),
                (11, int(sale.gross_profit))
            ]

            for col, value in money_cells:
                cell = ws.cell(row=row, column=col, value=value)
                cell.number_format = '#,##0'
            
            # % 데이터 작성
            percent_cell = ws.cell(row=row, column=12, value=profit_margin)
            percent_cell.number_format = '0.0%'
        
        # 열(col) 너비 조정
        column_widths = [8, 6, 15, 12, 15, 12, 12, 12, 12, 12, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # HTTP 응답 생성
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="월별 매출 현황.xlsx"'

        wb.save(response)
        return response
    
class IndividualMonthlySalesExcelExportView(LoginRequiredMixin, TemplateView):
    """
    월별 매출 상세 엑셀 파일 다운로드
    - 로그인한 사용자만 접근 가능
    - 특정 월별 매출 데이터를 엑셀 파일로 생성하여 다운로드
    """
    def get(self, request, *args, **kwargs):
        # url에서 pk 파라미터 추출
        sale_pk = kwargs.get('pk')
        try:
            sale = Sales.objects.get(pk=sale_pk)
        except Sales.DoesNotExist:
            messages.error(request, '해당 월별 매출이 존재하지 않습니다.')
            return redirect('sales_list')
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = f'{sale.year}년 {sale.month}월 매출'

        # 헤더 스타일 설정
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # 헤더 작성
        headers = [
            '연도', '월', '오프라인 매출', '기타 매출', '총 매출',
            '재료비', '노무비', '경비-소모품', '경비-기타', '총 원가', '매출이익', '수익률(%)', '메모'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 데이터 작성
        profit_margin = (sale.gross_profit / sale.total_sales) if sale.total_sales > 0 else 0

        # 기본 데이터 작성
        ws.cell(row=2, column=1, value=sale.year)
        ws.cell(row=2, column=2, value=sale.month)

        # 금액 데이터 작성
        money_cells = [
            (3, int(sale.offline_sales)),
            (4, int(sale.other_sales)),
            (5, int(sale.total_sales)),
            (6, int(sale.material_cost)),
            (7, int(sale.labor_cost)),
            (8, int(sale.supplies_expense)),
            (9, int(sale.other_expense)),
            (10, int(sale.total_cost)),
            (11, int(sale.gross_profit))
        ]

        for col, value in money_cells:
            cell = ws.cell(row=2, column=col, value=value)
            cell.number_format = '#,##0'

        # % 데이터 작성
        percent_cell = ws.cell(row=2, column=12, value=profit_margin)
        percent_cell.number_format = '0.0%'

        # 메모 작성
        ws.cell(row=2, column=13, value=sale.memo or '')

        # 열(col) 너비 조정
        column_widths = [8, 6, 15, 12, 15, 12, 12, 12, 12, 12, 15, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # HTTP 응답 생성
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{sale.year}년_{sale.month}월_매출.xlsx"'

        wb.save(response)
        return response

# Django Template Views for yearly sales management
class YearlySalesListView(LoginRequiredMixin, ListView):
    """
    연별 매출 목록 조회
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 연별 매출 목록 페이지를 렌더링
    """
    model = YearlySales
    template_name = 'glossymatcha/yearly_sales/list.html'
    context_object_name = 'yearly_sales_list'

class YearlySalesCreateView(LoginRequiredMixin, CreateView):
    """
    연별 매출 등록
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 연별 매출 등록 페이지를 렌더링
    """
    model = YearlySales
    form_class = YearlySalesForm
    template_name = 'glossymatcha/yearly_sales/create.html'
    success_url = reverse_lazy('yearly_sales_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.year}년 연별 매출이 등록되었습니다.')
        return response
    
class YearlySalesDetailView(LoginRequiredMixin, DetailView):
    """
    연별 매출 상세 정보
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 연별 매출 상세 페이지를 렌더링
    """
    model = YearlySales
    template_name = 'glossymatcha/yearly_sales/detail.html'
    context_object_name = 'yearly_sale'

class YearlySalesUpdateView(LoginRequiredMixin, UpdateView):
    """
    연별 매출 정보 수정
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 연별 매출 수정 페이지를 렌더링
    """
    model = YearlySales
    form_class = YearlySalesForm
    template_name = 'glossymatcha/yearly_sales/create.html'

    def get_success_url(self):
        return reverse_lazy('yearly_sales_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.year}년 연별 매출 정보가 수정되었습니다.')
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['yearly_sale'] = self.object
        return context
    
class YearlySalesDeleteView(LoginRequiredMixin, DeleteView):
    """
    연별 매출 삭제
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 연별 매출 삭제 페이지를 렌더링
    """
    model = YearlySales
    template_name = 'glossymatcha/yearly_sales/delete.html'
    success_url = reverse_lazy('yearly_sales_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(request, f'{self.object.year}년 연별 매출이 삭제되었습니다.')
        return super().delete(request, *args, **kwargs)
    
class YearlySalesExcelExportView(LoginRequiredMixin, TemplateView):
    """
    연별 매출 현황 엑셀 파일 다운로드
    - 로그인한 사용자만 접근 가능
    - 연별 매출 데이터를 엑셀 파일로 생성하여 다운로드
    """
    def get(self, request, *args, **kwargs):
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = '연별 매출 현황'

        # 헤더 스타일 설정
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # 콤마 포맷팅 설정
        number_style = NamedStyle(name='number_comma_yearly')
        number_style.number_format = '#,##0'

        # 퍼센트 포맷 설정
        percent_style = NamedStyle(name='percent_yearly')
        percent_style.number_format = '0.0%'

        # 헤더 작성
        headers = [
            '연도', '오프라인 매출', '기타 매출', '총 매출',
            '총 재료비', '총 노무비', '총 경비-소모품', '총 경비-기타', '총 원가', '매출 총이익', '수익률(%)'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 데이터 작성
        yearly_sales_data = YearlySales.objects.all().order_by('-year')
        
        for row, yearly_sale in enumerate(yearly_sales_data, 2):
            profit_margin = (yearly_sale.gross_profit / yearly_sale.total_sales) if yearly_sale.total_sales > 0 else 0

            # 기본 데이터 작성
            ws.cell(row=row, column=1, value=yearly_sale.year)

            # 금액 데이터 작성
            money_cells = [
                (2, int(yearly_sale.offline_sales)),
                (3, int(yearly_sale.other_sales)),
                (4, int(yearly_sale.total_sales)),
                (5, int(yearly_sale.total_material_cost)),
                (6, int(yearly_sale.total_labor_cost)),
                (7, int(yearly_sale.total_supplies_expense)),
                (8, int(yearly_sale.total_other_expense)),
                (9, int(yearly_sale.total_cost)),
                (10, int(yearly_sale.gross_profit))
            ]

            for col, value in money_cells:
                cell = ws.cell(row=row, column=col, value=value)
                cell.number_format = '#,##0'
            
            # % 데이터 작성
            percent_cell = ws.cell(row=row, column=11, value=profit_margin)
            percent_cell.number_format = '0.0%'
        
        # 열(col) 너비 조정
        column_widths = [8, 15, 12, 15, 12, 12, 12, 12, 12, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # HTTP 응답 생성
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="연별 매출 현황.xlsx"'

        wb.save(response)
        return response
    
class IndividualYearlySalesExcelExportView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        # url에서 pk 파라미터 추출
        yearly_sale_pk = kwargs.get('pk')
        try:
            yearly_sale = YearlySales.objects.get(pk=yearly_sale_pk)
        except YearlySales.DoesNotExist:
            messages.error(request, '해당 연별 매출이 존재하지 않습니다.')
            return redirect('yearly_sales_list')
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = f'{yearly_sale.year}년 매출'

        # 헤더 스타일 설정
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # 헤더 작성
        headers = [
            '연도', '오프라인 매출', '기타 매출', '총 매출',
            '총 재료비', '총 노무비', '총 경비-소모품', '총 경비-기타', '총 원가', '매출 총이익', '수익률(%)', '메모'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 데이터 작성
        profit_margin = (yearly_sale.gross_profit / yearly_sale.total_sales) if yearly_sale.total_sales > 0 else 0

        # 기본 데이터 작성
        ws.cell(row=2, column=1, value=yearly_sale.year)

        # 금액 데이터 작성
        money_cells = [
            (2, int(yearly_sale.offline_sales)),
            (3, int(yearly_sale.other_sales)),
            (4, int(yearly_sale.total_sales)),
            (5, int(yearly_sale.total_material_cost)),
            (6, int(yearly_sale.total_labor_cost)),
            (7, int(yearly_sale.total_supplies_expense)),
            (8, int(yearly_sale.total_other_expense)),
            (9, int(yearly_sale.total_cost)),
            (10, int(yearly_sale.gross_profit))
        ]

        for col, value in money_cells:
            cell = ws.cell(row=2, column=col, value=value)
            cell.number_format = '#,##0'

        # % 데이터 작성
        percent_cell = ws.cell(row=2, column=11, value=profit_margin)
        percent_cell.number_format = '0.0%'

        # 메모 작성
        ws.cell(row=2, column=12, value=yearly_sale.memo or '')

        # 열(col) 너비 조정
        column_widths = [8, 15, 12, 15, 12, 12, 12, 12, 12, 15, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # HTTP 응답 생성
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{yearly_sale.year}년_매출.xlsx"'

        wb.save(response)
        return response
    
@method_decorator(csrf_protect, name='dispatch')
class DailyPasswordCheckView(TemplateView):
    """
    일일 비밀번호 확인 페이지
    - 로그인한 사용자만 접근 가능
    - 인증된 경우 대시보드로 리다이렉트
    - 인증되지 않은 경우 인증 폼을 렌더링
    """
    template_name = 'glossymatcha/daily_password_check.html'

    def get(self, request, *args, **kwargs):
        """
        일일 비밀번호 확인 페이지
        - 이미 인증된 경우 대시보드로 리다이렉트
        - 인증되지 않은 경우 인증 폼을 렌더링
        """
        if request.session.get('daily_password_verified'):
            return redirect('dashboard')
        
        form = DailyPasswordForm()
        return render(request, self.template_name, {'form': form})
    
    def post(self, request, *args, **kwargs):
        form = DailyPasswordForm(request.POST)
        if form.is_valid():
            # 일일 패스워드 인증 성공
            request.session['daily_password_verified'] = True
            request.session['daily_password_date'] = str(date.today())
            messages.success(request, '일일 패스워드 인증이 완료되었습니다.')
            
            # Django 로그인 확인
            if not request.user.is_authenticated:
                return redirect('login')
            
            return redirect('dashboard')
        
        return render(request, self.template_name, {'form': form})
    
class DailyPasswordManagementView(LoginRequiredMixin, ListView):
    """
    일일 비밀번호 관리 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 일일 비밀번호 목록 페이지를 렌더링
    - 오늘 날짜의 비밀번호를 생성하고 컨텍스트에 추가
    """
    model = DailyPassword
    template_name = 'glossymatcha/daily_password/list.html'
    context_object_name = 'daily_passwords'

    def get_context_data(self, **kwargs):
        """
        일일 비밀번호 관리 페이지의 컨텍스트 데이터 추가
        - 오늘 날짜의 비밀번호를 생성하여 컨텍스트에 추가
        - 오늘 날짜의 비밀번호는 매일 새로 생성되며, 이전 비밀번호는 유지되지 않음
        - 오늘 날짜의 비밀번호는 DailyPassword 모델에서 생성된 값을 사용
        - 만약 오늘 날짜의 비밀번호가 없다면 새로 생성
        - 오늘 날짜의 비밀번호를 컨텍스트에 추가하여 템플릿에서 사용할 수 있도록 함
        - 이 컨텍스트는 템플릿에서 오늘 날짜의 비밀번호를 표시하는 데 사용됨
        - 이 컨텍스트는 일일 비밀번호 관리 페이지에서 오늘 날짜의 비밀번호를 표시하는 데 사용됨
        """
        context = super().get_context_data(**kwargs)
        today_password = DailyPassword.generate_today_password()
        context['today_password'] = today_password
        return context
    
class DailyPasswordCreateView(LoginRequiredMixin, CreateView):
    """
    일일 비밀번호 생성 페이지
    - 로그인한 사용자만 접근 가능
    - Django Template을 사용하여 일일 비밀번호 생성 페이지를 렌더링
    - 오늘 날짜의 비밀번호를 생성하고 성공 메시지를 표시
    """
    model = DailyPassword
    form_class = DailyPasswordManagementForm
    template_name = 'glossymatcha/daily_password/create.html'
    success_url = reverse_lazy('daily_password_management')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'{self.object.date} 일일 비밀번호가 생성되었습니다.')
        return response
    
def daily_password_logout(request):
    """
    일일 패스워드 인증 해제
    - 세션에서 인증 정보를 삭제하고 성공 메시지를 표시
    - 인증 해제 후 일일 패스워드 확인 페이지로 리다이렉트
    """
    if 'daily_password_verified' in request.session:
        del request.session['daily_password_verified']
    if 'daily_password_date' in request.session:
        del request.session['daily_password_date']
    messages.success(request, '일일 패스워드 인증이 해제되었습니다.')
    return redirect('daily_password_check')